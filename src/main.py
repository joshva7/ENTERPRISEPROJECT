from openpyxl import load_workbook
from validation import validate_candidate
from report_generator import save_report
from dashboard import show_dashboard

INPUT_FILE = "../input/Input_Data.xlsx"

workbook = load_workbook(INPUT_FILE)
sheet = workbook.active

shortlisted = []
rejected = []
errors = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    result = validate_candidate(row)

    if result["status"] == "Eligible":
        shortlisted.append(result)

    elif result["status"] == "Rejected":
        rejected.append(result)

    else:
        errors.append(result)

save_report(shortlisted, rejected, errors)
show_dashboard(shortlisted, rejected, errors)